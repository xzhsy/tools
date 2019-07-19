import openpyxl
import os


class FileInit(object):

    def __init__(self):
        path=os.path.join(os.getcwd(), 'demo.xlsx')
        self.wb = openpyxl.load_workbook('demo.xlsx')
        #print(self.wb.sheetnames)
        self.wd = self.wb['Sheet1']
        self.moduleen = []
        self.modulecn = []
        self.projectname = []
        self.sn = []
        self.level = []
        self.gitaddr = []

        self.m_list = []

    def gen_level(self):
        for i in range(2,self.wd.max_row+1):
            self.m_dict = {}
            #print(i)
            sn = self.wd.cell(row=i, column=1).value
            modulecn = self.wd.cell(row=i, column=2).value
            moduleen = self.wd.cell(row=i, column=3).value
            projectname = self.wd.cell(row=i, column=4).value
            level = self.wd.cell(row=i, column=5).value
            gitaddr = self.wd.cell(row=i, column=6).value
            self.m_dict['deploy_date'] = sn
            self.m_dict['business_name'] = modulecn
            self.m_dict['business_title'] = moduleen
            self.m_dict['deploy_order'] = level
            self.m_dict['git_address'] = gitaddr
            self.m_dict['project_title'] = projectname
            self.m_dict['deploy_date'] = sn
            #print(self.m_dict)

            self.m_list.append(self.m_dict)
        print(self.m_list)

if __name__ == '__main__':
    a= FileInit()
    a.gen_level()